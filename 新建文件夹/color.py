import os
import openpyxl
from openpyxl.styles import Color, Fill

wb = openpyxl.load_workbook('工作簿1.xlsx')
sheet = wb.get_active_sheet()

for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_column):
        cell = sheet.cell(row=i,column=j)
        print(cell.fill)
