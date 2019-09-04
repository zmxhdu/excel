# coding = utf-8
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def result_save(iCode, aType, mType, resultList, excelname, sheetname):
    result_df = []
    indexList_df = []
    if resultList['code'] == '-1':
        print(resultList['message'])
    else:
        count = 0
        for instrument, result in zip(instrumentList, resultList['result']):
            result_detail = []
            for instrument_key, instrument_value in instrument.items():
                if type(instrument_value) is list:
                    for i in range(0, len(instrument_value)):
                        if indexList_df is None:
                            index = instrument_key + '_' + str(i)
                            if count == 0:
                                indexList_df.append(index)
                            result_detail.append(instrument_value[i])
                else:
                    if count == 0:
                        indexList_df.append(instrument_key)
                    result_detail.append(instrument_value)

            for result_key, result_value in result.items():
                if type(result_value) is list:
                    for i in range(0, len(result_value)):
                        index = result_key + '_' + str(i)
                        if count == 0:
                            indexList_df.append(index)
                        result_detail.append(result_value[i])
                else:
                    if count == 0:
                        indexList_df.append(result_key)
                    result_detail.append(result_value)
            count += 1
            result_df.append(result_detail)

        result_data = pd.DataFrame(result_df, columns=indexList_df)

        try:
            wb = openpyxl.load_workbook(excelname)
            wb_sheet = wb.create_sheet(sheetname)
        except FileNotFoundError:
            wb = openpyxl.workbook.Workbook()
            wb_sheet = wb.get_active_sheet()
            wb_sheet.title = sheetname

        for r in dataframe_to_rows(result_data, index=False, header=True):
            wb_sheet.append(r)
        wb.save(excelname)


def batch_result_save(instrumentList, resultList, excelname, sheetname):
    result_df = []
    indexList_df = []
    if resultList['code'] == '-1':
        print(resultList['message'])
    else:
        count = 0
        for instrument, result in zip(instrumentList, resultList['result']):
            result_detail = []
            for instrument_key, instrument_value in instrument.items():
                if type(instrument_value) is list:
                    for i in range(0, len(instrument_value)):
                        if indexList_df is None:
                            index = instrument_key + '_' + str(i)
                            if count == 0:
                                indexList_df.append(index)
                            result_detail.append(instrument_value[i])
                else:
                    if count == 0:
                        indexList_df.append(instrument_key)
                    result_detail.append(instrument_value)

            for result_key, result_value in result.items():
                if type(result_value) is list:
                    for i in range(0, len(result_value)):
                        index = result_key + '_' + str(i)
                        if count == 0:
                            indexList_df.append(index)
                        result_detail.append(result_value[i])
                else:
                    if count == 0:
                        indexList_df.append(result_key)
                    result_detail.append(result_value)
            count += 1
            result_df.append(result_detail)

        result_data = pd.DataFrame(result_df, columns=indexList_df)

        try:
            wb = openpyxl.load_workbook(excelname)
            wb_sheet = wb.create_sheet(sheetname)
        except FileNotFoundError:
            wb = openpyxl.workbook.Workbook()
            wb_sheet = wb.get_active_sheet()
            wb_sheet.title = sheetname

        for r in dataframe_to_rows(result_data, index=False, header=True):
            wb_sheet.append(r)
        wb.save(excelname)
