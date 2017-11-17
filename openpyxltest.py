# coding=utf-8
import openpyxl
import re


wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_active_sheet()
# sheet页名称修改
sheet.title = 'Sheet1'
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        # 修改单元格样式为“常规”
        sheet.cell(row=i,column=j).number_format='General'
        c = sheet.cell(row=i, column=j)
        result = re.findall(r'(成\s+本)|(市\s+值)', str(c.value))
        if result:
            c.value = re.sub(r'(\s+)', '', str(c.value))
        print(c.value, c.number_format)

wb.save('test.xlsx')
