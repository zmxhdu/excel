# coding = utf-8
import openpyxl
import shutil
import os


try:
    xALR_property_rules_wb = openpyxl.load_workbook('银保监发[2018]28号附件2-财产险-校验样本.xlsx', read_only=False)  # 财产险校验规则
    xALR_life_rules_wb = openpyxl.load_workbook('银保监发[2018]28号附件3-人身险-校验样本.xlsx', read_only=False)  # 人身险校验规则
    xAlR_hide_rules_wb = openpyxl.load_workbook('xALR校验规则-隐藏规则.xlsx', read_only=False)  # 隐藏规则检查
except FileNotFoundError as rules_error:
    print(rules_error)
    exit()


# 需校验的财产险sheet名称列表
property_rules_sheets = ['表1-1 资产配置状况', '表1-2 资产信用状况', '表1-3 负债产品信息',
                         '表2-1 沉淀资金匹配（传统保险账户）', '表2-2 期限结构匹配（投资型和次级债账户）',
                         '表3-1 成本收益匹配状况', '表3-2 成本收益匹配压力测试',
                         '表4-1 现金流测试_普通账户', '表4-2 现金流测试_传统保险账户',
                         '表4-3 现金流测试_预定收益型投资保险产品账户', '表4-4 现金流测试_独立账户']

# 需校验的人身险sheet名称列表
life_rules_sheets = ['表1-1 资产配置状况', '表1-2 资产信用状况', '表1-3 负债产品信息',
                     '表2-1 期限结构匹配测试表_修正久期', '表2-2 期限结构匹配测试表_关键久期',
                     '表3-1 成本收益匹配状况表', '表3-2 成本收益匹配压力测试表',
                     '表4-1 现金流测试表_普通账户', '表4-2 现金流测试表_传统保险账户',
                     '表4-3 现金流测试表_分红保险账户', '表4-4 现金流测试表_万能保险账户', '表4-5 现金流测试表_独立账户']


# 判断报告类型是财产险还是人身险
def property_or_life_excle(file, filename):
    # TODO:增加判断“封面”是否存在
    sheet = filename['封面']

    if sheet['A6'].value == '财产保险公司资产负债管理量化评估表':
        print('待校验的报告类型是:', '财产险报告')
        xALR_property_rules_check(file, filename)
    elif sheet['A4'].value == '人身保险公司资产负债管理量化评估表':
        print('待校验的报告类型是:', '人身险报告')
        xALR_life_rules_check(file, filename)


# 财产险规则校验，蓝色单元格
def xALR_property_rules_check(file, filename):
    xALR_property_rules_sheets = xALR_property_rules_wb.sheetnames
    xALR_test_sheets = filename.sheetnames
    test_cell_error_count = 0

    if xALR_test_sheets != xALR_property_rules_sheets:
        print("待检验文件sheet页不符合要求")
        print("需要以下sheet页:", xALR_property_rules_sheets)
        print("带校验文件只有以下sheet页:", xALR_test_sheets)
    else:
        for sheet in property_rules_sheets:
            xALR_property_rules_sheet = xALR_property_rules_wb[sheet]
            xALR_test_sheet = filename[sheet]

            rows = xALR_property_rules_sheet.max_row
            cols = xALR_property_rules_sheet.max_column
            for row in range(1, rows+1):
                for col in range(1, cols+1):
                    cell = xALR_property_rules_sheet.cell(row, col)

                    # 公式+必须为空单元格校验
                    if cell.fill != None and cell.fill.start_color.rgb == 'FF00B0F0' or cell.fill.start_color.theme == 1:  # FF00B0F0 蓝色
                        test_cell = xALR_test_sheet.cell(row, col)
                        if not test_cell or test_cell.value != cell.value:
                            print(file, sheet, "单元格：", cell.coordinate, "不存在或公式错误", "正确公式为:", cell.value, "实际值为：", test_cell.value)
                            test_cell_error_count += 1

                    # 必填单元格校验
                    if cell.fill != None and cell.fill.start_color.theme == 7:
                        test_cell = xALR_test_sheet.cell(row, col)
                        if not test_cell or test_cell.value is None:
                            print(file, sheet, "单元格：", cell.coordinate, "数据漏填", "实际值为：", test_cell.value)
                            test_cell_error_count += 1

    # 隐藏规则校验，复制一份校验文件，并将结果输出到复制的文件中
    xALR_property_rules_wb_name, __ = os.path.basename(file).split('.')
    xALR_hide_rules_result = xALR_property_rules_wb_name + '_隐藏规则校验结果' + '.xlsx'
    shutil.copyfile(file, 'test.xlsx')
    shutil.copyfile(file, xALR_hide_rules_result)
    xALR_hide_rules_result_wb = openpyxl.load_workbook(xALR_hide_rules_result)
    xALR_hide_rules_result_sheet = xALR_hide_rules_result_wb.create_sheet(title='隐藏规则校验结果')

    xALR_hide_rules_result_sheet['A1'].value = '报表名称'
    xALR_hide_rules_result_sheet['B1'].value = '坐标'
    xALR_hide_rules_result_sheet['C1'].value = '坐标值'
    xALR_hide_rules_result_sheet['D1'].value = '规则1'
    xALR_hide_rules_result_sheet['E1'].value = '规则2'
    xALR_hide_rules_result_sheet['F1'].value = '规则3'
    xALR_hide_rules_result_sheet['G1'].value = '规则1校验结果'
    xALR_hide_rules_result_sheet['H1'].value = '规则2校验结果'
    xALR_hide_rules_result_sheet['I1'].value = '规则3校验结果'

    xALR_hide_rules_sheet = xAlR_hide_rules_wb['财产险'] # 个列说明  A:sheet页  B：坐标  D：校验规则
    xALR_hide_rules_sheet_rows = xALR_hide_rules_sheet.max_row
    xALR_hide_rules_sheet_cols = xALR_hide_rules_sheet.max_column

    # 加载规则，写入结果文件中
    for rows_a in range(2, xALR_hide_rules_sheet_rows):
        sheet_name = xALR_hide_rules_sheet['A%d' % rows_a].value  # 加载A列：报表名称
        coordinate = xALR_hide_rules_sheet['B%d' % rows_a].value  # 加载B列：坐标
        hide_rules = xALR_hide_rules_sheet['D%d' % rows_a].value  # 加载D列：隐藏规则

        xALR_hide_rules_result_sheet['A%d' % rows_a].value = sheet_name  # 写入A列：报表名称
        __, xALR_hide_rules_result_sheet['B%d' % rows_a].value = coordinate.split('!')  # 写入B列：坐标
        xALR_hide_rules_result_sheet['C%d' % rows_a].value = '=%s' % coordinate  # 写入C列：坐标值

        # 拆分规则
        hide_rule = hide_rules.split('|')
        for n in range(0, len(hide_rule)):
            # n+4：从D列开始写入拆分后的规则
            xALR_hide_rules_result_sheet.cell(rows_a, n + 4).value = '=%s' % hide_rule[n]
            rule_coordinate = xALR_hide_rules_result_sheet.cell(rows_a, n + 4).coordinate
            xALR_hide_rules_result_sheet.cell(rows_a, n + 7).value = '=C%d=%s' % (rows_a, rule_coordinate)

    xALR_hide_rules_result_wb.save(xALR_hide_rules_result)
    print('隐藏规则结果校验完成，结果见附件：', xALR_hide_rules_result)

    if test_cell_error_count > 0:
        print("存在 %d 处错误,校验不通过" % test_cell_error_count)
    else:
        print("校验通过")


# 人身险规则校验，黄色单元格
def xALR_life_rules_check(file, filename):
    xALR_life_rules_sheets = xALR_life_rules_wb.sheetnames
    xALR_test_sheets = filename.sheetnames
    test_cell_error_count = 0

    if xALR_test_sheets != xALR_life_rules_sheets:
        print("待检验文件sheet页不符合要求")
        print("需要以下sheet页:", xALR_life_rules_sheets)
        print("带校验文件只有以下sheet页:", xALR_test_sheets)
    else:
        for sheet in life_rules_sheets:
            xALR_lift_rules_sheet = xALR_life_rules_wb[sheet]
            xALR_test_sheet = filename[sheet]

            rows = xALR_lift_rules_sheet.max_row
            cols = xALR_lift_rules_sheet.max_column
            for row in range(1, rows+1):
                for col in range(1, cols+1):
                    cell = xALR_lift_rules_sheet.cell(row, col)

                    # 公式+必须为空单元格校验
                    if cell.fill != None and cell.fill.start_color.rgb == 'FFFFC000' or cell.fill.start_color.theme == 1:  # FFFFC000 黄色
                        test_cell = xALR_test_sheet.cell(row, col)
                        if not test_cell or test_cell.value != cell.value:
                            print(file, sheet, "单元格：", cell.coordinate, "不存在或公式错误", "正确公式为:", cell.value, "实际值为：", test_cell.value)
                            test_cell_error_count += 1

                    # 必填单元格校验
                    if cell.fill is not None and cell.fill.start_color.theme == 7:
                        test_cell = xALR_test_sheet.cell(row, col)
                        if not test_cell or test_cell.value != cell.value:
                            print(file, sheet, "单元格：", cell.coordinate, "数据漏填", "实际值为：", test_cell.value)
                            test_cell_error_count += 1

    # 隐藏规则校验，复制一份校验文件，并将结果输出到复制的文件中
    xALR_life_rules_wb_name, __ = os.path.basename(file).split('.')
    xALR_hide_rules_result = xALR_life_rules_wb_name + '_隐藏规则校验结果' + '.xlsx'
    shutil.copyfile(file, xALR_hide_rules_result)
    xALR_hide_rules_result_wb = openpyxl.load_workbook(xALR_hide_rules_result)
    xALR_hide_rules_result_sheet = xALR_hide_rules_result_wb.create_sheet(title='隐藏规则校验结果')

    xALR_hide_rules_result_sheet['A1'].value = '报表名称'
    xALR_hide_rules_result_sheet['B1'].value = '坐标'
    xALR_hide_rules_result_sheet['C1'].value = '坐标值'
    xALR_hide_rules_result_sheet['D1'].value = '规则1'
    xALR_hide_rules_result_sheet['E1'].value = '规则2'
    xALR_hide_rules_result_sheet['F1'].value = '规则3'
    xALR_hide_rules_result_sheet['G1'].value = '规则1校验结果'
    xALR_hide_rules_result_sheet['H1'].value = '规则2校验结果'
    xALR_hide_rules_result_sheet['I1'].value = '规则3校验结果'

    xALR_hide_rules_sheet = xAlR_hide_rules_wb['财产险']  # 个列说明  A:sheet页  B：坐标  D：校验规则
    xALR_hide_rules_sheet_rows = xALR_hide_rules_sheet.max_row
    xALR_hide_rules_sheet_cols = xALR_hide_rules_sheet.max_column

    # 加载规则，写入结果文件中
    for rows_a in range(2, xALR_hide_rules_sheet_rows):
        sheet_name = xALR_hide_rules_sheet['A%d' % rows_a].value  # 加载A列：报表名称
        coordinate = xALR_hide_rules_sheet['B%d' % rows_a].value  # 加载B列：坐标
        hide_rules = xALR_hide_rules_sheet['D%d' % rows_a].value  # 加载D列：隐藏规则

        xALR_hide_rules_result_sheet['A%d' % rows_a].value = sheet_name  # 写入A列：报表名称
        __, xALR_hide_rules_result_sheet['B%d' % rows_a].value = coordinate.split('!')  # 写入B列：坐标
        xALR_hide_rules_result_sheet['C%d' % rows_a].value = '=%s' % coordinate  # 写入C列：坐标值

        # 拆分规则
        hide_rule = hide_rules.split('|')
        for n in range(0, len(hide_rule)):
            # n+4：从D列开始写入拆分后的规则
            xALR_hide_rules_result_sheet.cell(rows_a, n + 4).value = '=%s' % hide_rule[n]
            rule_coordinate = xALR_hide_rules_result_sheet.cell(rows_a, n + 4).coordinate
            xALR_hide_rules_result_sheet.cell(rows_a, n + 7).value = '=C%d=%s' % (rows_a, rule_coordinate)

    xALR_hide_rules_result_wb.save(xALR_hide_rules_result)
    print('隐藏规则结果校验完成，结果见附件：', xALR_hide_rules_result)

    if test_cell_error_count > 0:
        print("存在 %d 处错误,校验不通过" % test_cell_error_count)
    else:
        print("校验通过")


if __name__ == '__main__':
    try:
        file = '银保监发(2018)28号附件2_安信0629-财产险 - 副本.xlsx'
        xALR_test_wb = openpyxl.load_workbook(file, read_only=False)
    except FileNotFoundError as file_error:
        print(file_error)
        exit()

    property_or_life_excle(file, xALR_test_wb)
