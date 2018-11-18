# coding=utf-8

from filecmp import *
import xlrd
from openpyxl import *
import re
import tkinter
import tkinter.filedialog, tkinter.messagebox, tkinter.scrolledtext
import sys


# xlsx格式文件对比
def compare_excel(ename1, ename2):

    print("------------------------------------")
    print("Comparing", ename1, ename2)
    # A bool to verify if 2 xlsx is the same.
    fileSame = True

    # Load workbook, and get sheetname
    wb1 = load_workbook(filename=ename1)
    wb2 = load_workbook(filename=ename2)
    sn1 = wb1.sheetnames
    sn2 = wb2.sheetnames

    if sn1 != sn2:
        print("Two file has different sheets.")
        print(ename1, "has sheet names:", sn1)
        print(ename2, "has sheet names:", sn2)
    else:
        sn = sn1
        for wsn in sn:
            # Get worksheet
            ws1 = wb1[wsn]
            ws2 = wb2[wsn]
            c = ws1.max_column
            r = ws1.max_row
            # This can be replace by ws1.get_dimension
            if (ws2.max_column != c) or (ws2.max_row != r):
                print("DIFFERDENT at SHEET-", wsn, ": Rows or columns not the same!")
                fileSame = False
            else:
                # Compare every cell.
                flag = True
                for i in range(1, r+1):
                    for j in range(1, c+1):
                        c1 = ws1.cell(i, j)
                        c2 = ws2.cell(i, j)
                        if c1:
                            if c2:
                                if c1.value != c2.value:
                                    if (wsn == "Internal Info") and ((i == 4) and (j == 2)) or ((i == 5) and (j == 3)):
                                        continue
                                    print("DIFFERDENT_VALUE at SHEET-", wsn, ": At (", i, ",", j, ")", )
                                    print("diff FROM", c1.value, "TO", c2.value)
                                    flag = False
                            else:
                                print("DIFFERDENT_TO_NONE at SHEET-", wsn, ": At (", i, ",", j, ")")
                                print("diff FROM", c1.value)
                                flag = False
                        else:
                            if c2:
                                print("DIFFERDENT_TO_NONE at SHEET-", wsn, ": At (", i, ",", j, ")")
                                print("diff FROM", c2.value)
                                flag = False
                fileSame = fileSame and flag
        if fileSame:
            print("SAME_FILE:", ename1, ename2)

    print("------------------------------------")
    print()


# 对比xls格式文件
def read_excel(ename1, ename2):
    print("------------------------------------")
    print("Comparing", ename1, ename2)
    # A bool to verify if 2 xlsx is the same.
    fileSame = True

    # Load workbook, and get sheetname
    wb1 = xlrd.open_workbook(filename=ename1)
    wb2 = xlrd.open_workbook(filename=ename2)
    sn1 = wb1.sheet_names()
    sn2 = wb2.sheet_names()

    if sn1 != sn2:
        print("Two file has different sheets.")
        print(ename1, "has sheet names:", sn1)
        print(ename2, "has sheet names:", sn2)
    else:
        sn = sn1
        for wsn in sn:
            # Get worksheet
            ws1 = wb1.sheet_by_name(wsn)
            ws2 = wb2.sheet_by_name(wsn)
            c = ws1.ncols
            r = ws1.nrows
            # This can be replace by ws1.get_dimension
            if (ws2.ncols != c) or (ws2.nrows != r):
                print("DIFFERDENT at SHEET-", wsn, ": Rows or columns not the same!")
                fileSame = False
            else:
                # Compare every cell.
                flag = True
                for i in range(0, r):
                    for j in range(0, c):
                        c1 = ws1.cell(i, j)
                        c2 = ws2.cell(i, j)
                        if c1:
                            if c2:
                                if c1.value != c2.value:
                                    if (wsn == "Internal Info") and ((i == 4) and (j == 2)) or (
                                            (i == 5) and (j == 3)):
                                        continue
                                    print("DIFFERDENT_VALUE at SHEET-", wsn, ": At (", i, ",", j, ")", )
                                    print("diff FROM", c1.value, "TO", c2.value)
                                    flag = False
                            else:
                                print("DIFFERDENT_TO_NONE at SHEET-", wsn, ": At (", i, ",", j, ")")
                                print("diff FROM", c1.value)
                                flag = False
                        else:
                            if c2:
                                print("DIFFERDENT_TO_NONE at SHEET-", wsn, ": At (", i, ",", j, ")")
                                print("diff FROM", c2.value)
                                flag = False
                fileSame = fileSame and flag
        if fileSame:
            print("SAME_FILE:", ename1, ename2)

    print("------------------------------------")
    print()


# 对比目录
# dirName = '情景分析'
# dirFB011 = 'E:\测试用例_zmx\版本用例/4.2.0.0200\\fb02发布测试\第二次对比\FB01\风险分析' + '\\' + dirName
# dirFB012 = 'E:\测试用例_zmx\版本用例/4.2.0.0200\\fb02发布测试\第二次对比\FB02\风险分析' + '\\' + dirName

def test_compare():
    try:
        compare_result = open(result_file, 'a')
        sys.stdout = compare_result

        a = dircmp(dir_old, dir_new)
        a.report_full_closure()

        for ex in a.diff_files:
            if re.search('.xlsx$', str(ex)) is not None:
                excl1 = str(dir_old) + '\\' + str(ex)
                excl2 = str(dir_new) + '\\' + str(ex)
                compare_excel(excl1, excl2)
            elif re.search('.xls$', str(ex)) is not None:
                excl1 = str(dir_old) + '\\' + str(ex)
                excl2 = str(dir_new) + '\\' + str(ex)
                read_excel(excl1, excl2)

        print("------------------------------------")
        compare_result.close()

        tkinter.messagebox.showinfo('提示', '对比结束')
        result_txt = tkinter.scrolledtext.ScrolledText(root, width=50)
        result_txt.grid(row=6, column=1)

        with open(result_file, 'r') as f:
            for readline in f:
                result_txt.insert(tkinter.INSERT, readline)
    except NameError:
        tkinter.messagebox.showerror('错误','文件路径异常，请重新选择')



def SelectPathOld():
    path_ = tkinter.filedialog.askdirectory()
    global dir_old
    dir_old = path_
    path_old.set(path_)


def SelectPathNew():
    path_ = tkinter.filedialog.askdirectory()
    global dir_new
    dir_new = path_
    path_new.set(path_)


def SelectPathResult():
    path_ = tkinter.filedialog.askopenfilename()
    global result_file
    result_file = path_
    path_result.set(path_)


root = tkinter.Tk()
root.title("发布测试对比")
root.geometry("550x450")

path_old = tkinter.StringVar()
path_new = tkinter.StringVar()
path_result = tkinter.StringVar()


l1 = tkinter.Label(root, text="上版本路径")
l1.grid(row=0, column=0)
dirold_entry = tkinter.Entry(root, width=50, textvariable=path_old)
dirold_btn = tkinter.Button(root, text="路径选择", command=SelectPathOld)
dirold_entry.grid(row=0, column=1)
dirold_btn.grid(row=0, column=2)


l2 = tkinter.Label(root, text="发布版本路径")
l2.grid(row=2, column=0)
dirnew_entry = tkinter.Entry(root, width=50, textvariable=path_new)
dirnew_btn = tkinter.Button(root, text="路径选择", command=SelectPathNew)
dirnew_entry.grid(row=2, column=1)
dirnew_btn.grid(row=2, column=2)

l3 = tkinter.Label(root, text="对比结果输出文件")
l3.grid(row=4, column=0)
result_entry = tkinter.Entry(root, width=50, textvariable=path_result)
result_btn = tkinter.Button(root, text="路径选择", command=SelectPathResult)
result_entry.grid(row=4, column=1)
result_btn.grid(row=4, column=2)

compare_button = tkinter.Button(root, text="开始对比", command=test_compare)
compare_button.grid(row=5, column=1)

result = tkinter.Label(root, text="结果预览")
result.grid(row=6,column=0)

root.mainloop()
